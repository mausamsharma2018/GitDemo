import openpyxl

book = openpyxl.load_workbook("C:\\Users\\susil\\Mausam\\pyexcel.xlsx")
sheet = book.active
Dict={}
cell = sheet.cell(row=2, column=2)
#printing the cell value.
print(cell.value)

#updating the value in the excel sheet
sheet.cell(row=2,column=1).value= "mausam"
# priitng the value to check if the value got updated
print(sheet.cell(row=2,column=1).value)

#to print row number
print(sheet.max_row)
#to print column numbers
print(sheet.max_column)

print(sheet['A2'].value) # from column a and row 2.

#printing all the value from the excel sheet

for i in range(1,sheet.max_row+1): # to get rows
    if sheet.cell(row=i,column=1).value=="Delivery Truck": # if want to get data only for the particular testcase.write testcase name instead of delivery truck in "".
       for j in range(1,sheet.max_column+1): # to get columns
          print(sheet.cell(row=i,column=j).value)
          #Dict["firstname"]="rahul"
          #insertingt he values in one dictionary ,so that we can create the test data utility from this .
          Dict[sheet.cell(row=2,column=j).value]=sheet.cell(row= i, column=j).value

print(Dict)